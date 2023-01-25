import base64
import csv
import datetime
import io
import math
import os
import re
import time

import ciso8601 as ciso8601
import dateutil
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import pdfkit
import doctest
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Font, Side, Border
from prettytable import PrettyTable, ALL

dic_transl = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Нижняя граница вилки оклада',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary': 'Оклад',
    'salary_average': 'Средний оклад',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
    '№': '№',
    'noExperience': 'Нет опыта',
    'between1And3': 'От 1 года до 3 лет',
    'between3And6': 'От 3 до 6 лет',
    'moreThan6': 'Более 6 лет',
    'False': 'Нет',
    'True': 'Да',
    'AZN': 'Манаты',
    'BYR': 'Белорусские рубли',
    'EUR': 'Евро',
    'GEL': 'Грузинский лари',
    'KGS': 'Киргизский сом',
    'KZT': 'Тенге',
    'RUR': 'Рубли',
    'UAH': 'Гривны',
    'USD': 'Доллары',
    'UZS': 'Узбекский сум'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

vacancies_name = {'Backend-программист ': ['backend', 'бэкэнд', 'бэкенд', 'бекенд', 'бекэнд', 'back end', 'бэк энд',
                                           'бэк енд', 'django', 'flask', 'laravel', 'yii', 'symfony'],'Аналитик': ['analytic', 'аналитик', 'analyst', 'аналітик'],
                  '1С-разработчик': ['1с разработчик', '1c разработчик', '1с', '1c', '1 c', '1 с'],
                  'Руководитель ИТ-проектов': ['team lead', 'тимлид', 'тим лид', 'teamlead', 'lead', 'руководит',
                                               'директор', 'leader', 'director', 'начальник', 'лидер',
                                               'управляющий проект', 'керівник', 'chief', 'начальник it'],
                  'Специалист техподдержки': ['техподдержка', 'тех поддержка', 'technical support engineer',
                                              'поддержка', 'support', 'підтримки'],
                  'Системный администратор': ['system admin', 'сисадмин', 'сис админ', 'системный админ',
                                              'cистемный админ', 'администратор систем', 'системний адміністратор'],
                  'UX/UI дизайнер': ['design', 'ux', 'ui', 'дизайн', 'иллюстратор'],
                  'Менеджер IT-проекта': ['project manager', 'менеджер проект', 'менеджер it проект',
                                          'менеджер ит проект', 'менеджер интернет проект', 'проджект менеджер',
                                          'проект менеджер', 'проектный менеджер', 'менеджер по проект',
                                          'менеджер по сопровождению проект', 'управление проект',
                                          'управлению проект',
                                          'project менедж', 'администратор проект', 'менеджер проектів',
                                          'менеджер it продукт', 'менеджер it product'],
                  'Тестировщик (QA-инженер)': ['qa', 'test', 'тест', 'quality assurance'],
                  'Инженер-программис': ['engineer', 'инженер программист', 'інженер', 'it инженер',
                                         'инженер разработчик'],
                  'Frontend-программист ': ['frontend', 'фронтенд', 'вёрстка', 'верстка', 'верста', 'front end',
                                            'angular', 'html', 'css', 'react', 'vue'],
                  'Специалист по информационной безопасности': ['безопасность', 'защита',
                                                                'information security specialist',
                                                                'information security', 'фахівець служби безпеки',
                                                                'cyber security'],
                  'ERP-специалист': ['erp', 'enterprise resource planning', 'abap', 'crm', 'help desk', 'helpdesk',
                                     'service desk', 'servicedesk', 'bi', 'sap'],
                  'Backend-программист ': ['backend', 'бэкэнд', 'бэкенд', 'бекенд', 'бекэнд', 'back end', 'бэк энд',
                                           'бэк енд', 'django', 'flask', 'laravel', 'yii', 'symfony'],
                  'Java-программист': ['java', 'ява', 'джава'],
                  'Администратор баз данных ': ['баз данных', 'оператор баз данных', 'базы данных', 'oracle',
                                                'mysql',
                                                'data base', 'database', 'dba', 'bd', 'бд', 'базами данны'],
                  'Devops-инженер ': ['devops', 'development operations'], 'PHP-программист': ['php', 'пхп', 'рнр'],
                  'Web-разработчик ': ['web develop', 'веб разработчик', 'web разработчик', 'web programmer',
                                       'web программист', 'веб программист', 'битрикс разработчик',
                                       'bitrix разработчик', 'drupal разработчик', 'cms разработчик',
                                       'wordpress разработчик', 'wp разработчик', 'joomla разработчик',
                                       'drupal developer', 'cms developer', 'wordpress developer', 'wp developer',
                                       'joomla developer'], 'Python-программист': ['python', 'питон', 'пайтон'],
                  'C/C++ программист': ['c\+\+', 'с\+\+'],
                  'Android-разработчик ': ['android', 'андроид', 'andorid', 'andoroid', 'andriod', 'andrind',
                                           'xamarin'],
                  'Разработчик игр (GameDev)': ['game', 'unity', 'игр', 'unreal'],
                  'Fullstack-программист': ['fullstack', 'фулстак', 'фуллтак', 'фуллстэк', 'фулстэк', 'full stack'],
                  'IOS-разработчик ': ['ios'], 'C# программист': ['c#', 'c sharp', 'шарп', 'с#']
                  }


class DataSet:
    """
    Класс для работы с файлом и создания колекции объектов Vacancy

    Atribyties:
        file_name (str): Имя файла для считывания
        vacancies_objects (list of Vacancy): лист вакансий

    >>> type(DataSet('F:/Makarov/small.csv')).__name__
    'DataSet'
    """

    def __init__(self, file_name=None):
        """
        Иницилизирует объект DataSet
        Args:
            file_name (str): Имя файла для считывания
        vacancies_objects (list of Vacancy): лист вакансий
        """
        self.file_name = file_name
        if file_name != None:
            self.vacancies_objects = self.parser_csv(file_name)
        else:
            self.vacancies_objects = None

    def __add__(self, data):
        return self.setvacancies(self.getvacancies() + data.getvacancies())

    def setvacancies(self, vacancies_objects):
        """
        Перазаписывает лист вакансий
        vacancies_objects(list of Vacancy): лист вакансий
        """
        self.vacancies_objects = vacancies_objects

    def getvacancies(self):
        """
        Возврашает лист вакагсий из объкта DataSet
        Returns:
            list of Vacancy: лист вакансий
        """
        return self.vacancies_objects

    @staticmethod
    def parser_csv(name):
        """
        Читает csv файл и создает лист объектов Vacancy
        Args:
            name (str): Имя файла для считывания из файла
        Returns:
            list of Vacancy: лист вакансий
        """
        if os.stat(name).st_size == 0:
            return []
        flag = True
        counter = 0
        data = []
        with open(name, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            heading = next(reader)
            for row in reader:
                for piece in row:
                    if len(piece) == 0:
                        flag = False
                    counter += 1
                if flag and (len(heading) == counter):
                    inf = DataSet.clear_row_csv(row, heading)
                    for key in ['name', 'description', 'key_skills', 'experience_id', 'premium', 'employer_name',
                                'salary', 'area_name',
                                'published_at']:
                        if not inf.__contains__(key):
                            inf[key] = None
                    data.append(Vacancy(inf['name'],
                                        inf['description'],
                                        inf['key_skills'],
                                        inf['experience_id'],
                                        inf['premium'],
                                        inf['employer_name'],
                                        inf['salary'],
                                        inf['area_name'],
                                        inf['published_at']))
                flag = True
                counter = 0
        return data

    # @staticmethod
    # def format_time(date_str):
    #     return datetime.datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S%z')
    #
    # @staticmethod
    # def format_time1(date_str):
    #     return datetime.datetime(int(date_str[:4]), int(date_str[5:7]),
    #                              int(date_str[8:10]),
    #                              int(date_str[11:13]), int(date_str[14:16]),
    #                              int(date_str[17:19]), tzinfo=datetime.timezone(
    #             datetime.timedelta(seconds=int(date_str[21:]) * 36)))
    #
    # @staticmethod
    # def format_time2(date_str):
    #     return dateutil.parser.isoparse(date_str)

    @staticmethod
    def format_time3(date_str):
        """
        Переводит вермя из строки в DateTime
        Args:
            row(str): дата в виде строки'
        Returns:
            datetime: время в вормате DateTime
        """
        return ciso8601.parse_datetime(date_str)

    @staticmethod
    def formatter(row):
        """
        Получает данные о вакансии и переводит их в нужный формат
        Args:
            row(dict): словарь с данными о вакансии
        Returns:
            dict: изменнеый с данными о вакансии
        """
        dict_new = {'name': row['name']}
        try:
            dict_new = {'description': row['description'],
                        'key_skills': ('\n'.join(row['key_skills'])),
                        'experience_id': dic_transl[row['experience_id']],
                        'premium': 'Да' if (row['premium'] == 'True') else 'Нет',
                        'employer_name': row['employer_name']}
            salary_from = "{0:,}".format((int(float(row['salary_from'])))).replace(',', ' ')
            salary_to = "{0:,}".format((int(float(row['salary_to'])))).replace(',', ' ')
            salary_text = f"{salary_from} - {salary_to} ({dic_transl[row['salary_currency']]}) ({'С вычетом' if row['salary_gross'] == 'False' else 'Без вычета'} налогов)"
        except Exception:
            salary_text = None
            row['salary_gross'] = None
        dict_new['salary'] = Salary(salary_text, row["salary_from"], row["salary_to"], row['salary_gross'],
                                    row['salary_currency'], math.floor(
                (int(row['salary_from'].replace('.0', '')) + int(row['salary_to'].replace('.0', ''))) / 2))
        dict_new['area_name'] = row['area_name'] if row.__contains__('area_name') else None
        dict_new['published_at'] = DataSet.format_time3(row['published_at'])
        return dict_new

    @staticmethod
    def delete_tags(row):
        temp_value = ''
        while row.find('<') != -1:
            temp_value += row[:row.find('<')]
            current_index = row.find('>') + 1
            row = row[current_index:]
        else:
            return temp_value + row

    @staticmethod
    def clear_row_csv(row, list_naming):
        """
        Чистит вакансию от тегов и вызывает метод форматирования вакансии
        Args:
            row(list of str): лист с данными
            list_naming(list of str):лист с названиями
        Returns:
            dict: словарь с очишиными данными о вакансии
        """
        dict_new = {}
        # pattern = re.compile('<.*?>')
        # pattern1 = re.compile('\s+')
        for i in range(0, len(list_naming)):
            if list_naming[i] == 'key_skills':
                dict_new[list_naming[i]] = row[i].split("\n")
            else:
                row[i] = DataSet.delete_tags(row[i])
                row[i] = row[i].replace('  ', ' ')
                # row[i] = row[i].replace("\n", ", ")
                # row[i] = re.sub(pattern, '', row[i])
                # row[i] = re.sub(pattern1, ' ', row[i])
                row[i] = row[i].strip()
                dict_new[list_naming[i]] = row[i]

        return DataSet.formatter(dict_new)


class Vacancy:
    """
    Класс для удобной работы с вакансиями

    Atribyties:
        name (str): название вакансии
        description(str): описание вакансии
        key_skills(str): навыки необходимые для ваканцсии (разделены \n)
        experience_id(str): необходимый опыт для вакансии
        premium(str): 'Да' если вакансия перммиум и 'Нет' если нет
        employer_name(str): название организации
        salary (Salary): информация о зарплатах
        area_name(str): информация о месте5 находждения вакансии
        published_at(datetime): неотформатированнная дата (используятся сравнения)


    """

    def __init__(self, name, description, key_skills, experience_id, premium, employer_name, salary, area_name,
                 published_at):
        """
        Иницилизация объекта Vacancy
        Args:
            name (str): название вакансии
            description(str): описание вакансии
            key_skills(str): навыки необходимые для ваканцсии (разделены \n)
            experience_id(str): необходимый опыт для вакансии
            premium(str): 'Да' если вакансия перммиум и 'Нет' если нет
            employer_name(str): название организации
            salary (Salary): информация о зарплатах
            area_name(str): информация о месте5 находждения вакансии
            published_at(datetime): неотформатированнная дата (используятся сравнения)
        """
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

    def getdata(self):
        """
        Возврашает лист с данными о вакансии
        Returns:
            list: данные о вакансии

        """
        return [self.name, self.description, self.key_skills, self.experience_id, self.premium, self.employer_name,
                self.salary.salary_text, self.area_name, self.published_at]


class Salary:
    """
    Класс для работы с зарплатами в объкте Vacancy
    Atribyties:
        salary_text(str): Текст с данными о зарплате
        salary_from(float): Максимальная зарплата
        salary_to(float): Минимальная зарплата
        salary_gross(bool): с вычетом налогов или без
        salary_currency(str): валюта
        salary_average(float):среднее значение зарплаты
    """

    def __init__(self, salary_text, salary_from, salary_to, salary_gross, salary_currency, salary_average):
        """
        Иницилизирует объект Salary
        Args:
            salary_text(str): Текст с данными о зарплате
            salary_from(float): Максимальная зарплата
            salary_to(float): Минимальная зарплата
            salary_gross(bool): с вычетом налогов или без
            salary_currency(str): валюта
            salary_average(float):среднее значение зарплаты
        """
        self.salary_text = salary_text
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency
        self.salary_average = salary_average


class InputConnect:
    """
    Класс созданный для работы с таблицами и параметрами для её создания
    """

    @staticmethod
    def check_paramtrs(parameters):
        """
        Проверяет параметры сортировки
        Args:
            parameters(str): параметры для формирования таблицы
        Returns:
            bool: верны ли параметры сортировки
        """
        if len(parameters) == 0:
            return True
        parameters_list = parameters.split(': ')
        if ':' not in parameters:
            print('Формат ввода некорректен')
            return False
        if parameters[:parameters.index(':')] not in dic_transl.values():
            print('Параметр поиска некорректен')
            return False
        parameters_list[0] = ''.join(filter(lambda x: parameters_list[0] == dic_transl[x], dic_transl))
        return bool(parameters_list[0])

    @staticmethod
    def check_sortparam(sort_info):
        """
        Проверяет параметры форматирования таблицы
        Args:
            sort_info(list of str): параметры для форматирования таблицы
        Returns:
             bool: верны ли параметры форматирования таблицы
        """
        if (sort_info[0] not in dic_transl.values()) and (sort_info[0] != ''):
            print('Параметр сортировки некорректен')
            return False
        if (sort_info[1] != 'Да') and (sort_info[1] != 'Нет') and (sort_info[1] != ''):
            print('Порядок сортировки задан некорректно')
            return False
        return True

    @staticmethod
    def sort_vacancies(vacancies, sort_info):
        """
        Сортирует вакансии по указаным параметрам
        Ags:
            vacancies(list of Vacancy): лист объектов Vacancy
            sort_info(list of str): Данные для сортировки
        Returns:
            list of Vacancy: отсортирвоанные лист вакансий
        """
        if not InputConnect.check_sortparam(sort_info):
            return vacancies
        vacancies_sort = vacancies
        if sort_info[0] == 'Оклад':
            data_sort = sorted(vacancies_sort,
                               key=lambda x: x.salary.salary_average * currency_to_rub[x.salary.salary_currency],
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Навыки':
            data_sort = sorted(vacancies_sort, key=lambda x: x.key_skills.count('\n'),
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Опыт работы':
            experience_weight = {'Нет опыта': 0, 'От 1 года до 3 лет': 1, 'От 3 до 6 лет': 2, 'Более 6 лет': 3}
            data_sort = sorted(vacancies_sort, key=lambda x: experience_weight[x.experience_id],
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Дата публикации вакансии':
            data_sort = sorted(vacancies_sort, key=lambda x: x.published_at,
                               reverse=True if sort_info[1] == 'Да' else False)
        else:
            data_sort = sorted(vacancies_sort,
                               key=lambda x: getattr(x, ''.join(
                                   filter(lambda x: sort_info[0] == dic_transl[x], dic_transl))),
                               reverse=True if sort_info[1] == 'Да' else False)
        return data_sort

    @staticmethod
    def data_filter(vacancies, parameters, sort_info):
        """
        Фильтрует вакансии по заданным параметрам
        Argumenst:
            vacancies(list of Vacancy): квакансии для сортировки
            parameters(str): параметры для формирования таблицы
            sort_info(list of str): Данные для сортировки
        Returns:
            list of Vacancy: отфильтрованый лист вакансий
        """
        if len(parameters) == 0:
            if sort_info[0] != '':
                return InputConnect.sort_vacancies(vacancies, sort_info)
            return vacancies
        parameters_list = parameters.split(': ')
        data = []
        parameters_list[0] = ''.join(filter(lambda x: parameters_list[0] == dic_transl[x], dic_transl))
        i = ''.join(filter(lambda x: parameters_list[1] == dic_transl[x], dic_transl))
        if i and parameters_list[0] not in ['experience_id', 'premium']:
            parameters_list[1] = i
        if parameters_list[0] == 'key_skills':
            parameters_list[1] = parameters_list[1].split(', ')
            for vacancy in vacancies:
                f = True
                for parametr in parameters_list[1]:
                    if (parametr + '\n') not in vacancy.key_skills:
                        f = False
                if f:
                    data.append(vacancy)
        elif parameters_list[0] == 'salary':
            for vacancy in vacancies:
                if float(vacancy.salary.salary_from) <= float(parameters_list[1]) <= float(vacancy.salary.salary_to):
                    data.append(vacancy)
        elif parameters_list[0] == 'published_at':
            for vacancy in vacancies:
                if vacancy.published_at.strftime('%d.%m.%Y') == parameters_list[1]:
                    data.append(vacancy)
        elif parameters_list[0] == 'salary_currency':
            for vacancy in vacancies:
                if parameters_list[1] == vacancy.salary.salary_currency:
                    data.append(vacancy)
        else:
            for vacancy in vacancies:
                if parameters_list[1] == getattr(vacancy, parameters_list[0]):
                    data.append(vacancy)
        if sort_info[0] != '':
            data = InputConnect.sort_vacancies(data, sort_info)
        return data

    @staticmethod
    def cut_table(table, num_str, name_colm):
        """
        Обрезает созданую таблицу по перадным параметрам
        Arguments:
            table(PrettyTable): таблица
            num_str: номера строк
            name_colm: имена колонок
        Returns:
            PrettyTable: таблица
        """
        if len(num_str) == 1:
            return table.get_string(start=0 if num_str[0] == '' else int(num_str[0]) - 1,
                                    fields=table.field_names if name_colm[0] == '' else ['№'] + name_colm)
        else:
            return table.get_string(start=int(num_str[0]) - 1, end=int(num_str[1]) - 1,
                                    fields=table.field_names if name_colm[0] == '' else ['№'] + name_colm)

    @staticmethod
    def translate(list_naming):
        """
        Перовдит слова через словарь
        Arguments:
            list_naming(list of str): слова для перевода
        Returns:
            list of str: переведенные слова
        """
        list_translate = []
        for name in list_naming:
            list_translate.append(dic_transl[name])
        return list_translate

    @staticmethod
    def create_table(vacancies):
        """
        Формирует таблицу через класс PretyTable
        Arguments:
            vacancies(list of Vacancy): лист вакансий
        Returns:
             PretyTabel: таблица
        """
        data = []
        for vacancy in vacancies:
            data.append(vacancy)
        table = PrettyTable()
        table.field_names = ['№'] + InputConnect.translate(list(data[0].__dict__))
        table.hrules = ALL
        table.align = 'l'
        table.max_width = 20
        for i, row in enumerate(data):
            try:
                # row['key_skills'] = reduce(lambda x, y: x + '\n' + y, row['key_skills'])
                if len(row.key_skills) > 100:
                    row.key_skills = row.key_skills[:100] + '...'
                if len(row.description) > 100:
                    row.description = row.description[:100] + '...'
            except Exception:
                pass
            if (i > 100_000):
                print('Файл слишком большой, могу обработать только часть')
                return table
            table.add_row([i + 1] + row.getdata())

        return table


class Statistics:
    """
    Класс для создания статистики по вакансии и ее сравнения
    с общей статистикой
    Atribyties:
        dataset(DataSet): данные для формирования статистики
        dict_salary(dict{int:float}): статистика год:средння зарплата
        dict_quantity(dict{int:int}): статистика год:количество вакантных мест
        dict_salary_name(dict{int:float}): статистика год:средння зарплата для выбранной вакансии
        dict_quantity_name(dict{int:int}): статистика год:количество вакантных мест для выбранной вакансии
        dict_salary_city(dict{str:float}): статистика город:средння зарплата (не более 10)
        dict_vacancy_share(dict{str:float}): статистика город:относительное число вакансий (не более 10)
        dict_salary_city_name(dict{str:float}): статистика город:средння зарплата   для выбранной вакансии(не более 10)
        dict_vacancy_share_name(dict{str:float}): статистика город:относительное число вакансий   для выбранной вакансии(не более 10)
    """

    def __init__(self, dataset):
        """
        Иницилизация объекта класс Salary
        Args:
            dataset(DataSet): данные для формирования статистики
        """
        self.dataset = dataset
        self.dict_salary = {}
        self.dict_quantity = {}
        self.dict_salary_name = {}
        self.dict_quantity_name = {}
        self.dict_salary_city = {}
        self.dict_vacancy_share = {}
        self.dict_salary_city_name = {}
        self.dict_vacancy_share_name = {}

    def __filling_dict(self, name):
        """
        Заполняет словари со статистиками
        Args:
            name(str): название вакансии
        """

        # chk_pat = '(?:{})'.format('|'.join(['техподдержка', 'тех поддержка', 'technical support engineer', 'поддержка', 'support', 'підтримки']))
        chk_pat = '(?:{})'.format('|'.join(vacancies_name[name]))
        for vacancy in self.dataset.getvacancies():
            if not self.dict_salary.__contains__(vacancy.published_at.year):
                self.dict_salary[vacancy.published_at.year] = vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_quantity[vacancy.published_at.year] = 1
            else:
                self.dict_salary[vacancy.published_at.year] += vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_quantity[vacancy.published_at.year] += 1

            if not self.dict_salary_city.__contains__(vacancy.area_name):
                self.dict_salary_city[vacancy.area_name] = vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_vacancy_share[vacancy.area_name] = 1
            else:
                self.dict_salary_city[vacancy.area_name] += vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_vacancy_share[vacancy.area_name] += 1

            if re.search(chk_pat, vacancy.name.lower(), flags=re.I):
                if not self.dict_salary_city_name.__contains__(vacancy.area_name):
                    self.dict_salary_city_name[vacancy.area_name] = vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_vacancy_share_name[vacancy.area_name] = 1
                else:
                    self.dict_salary_city_name[vacancy.area_name] += vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_vacancy_share_name[vacancy.area_name] += 1

                if not self.dict_salary_name.__contains__(vacancy.published_at.year):
                    self.dict_salary_name[vacancy.published_at.year] = vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_quantity_name[vacancy.published_at.year] = 1
                else:
                    self.dict_salary_name[vacancy.published_at.year] += vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_quantity_name[vacancy.published_at.year] += 1
    def __sort_dict(self):
        """
        Сортирует словари с вакансиями
        """
        for item in self.dict_salary.keys():
            try:
                self.dict_salary[item] = math.trunc(self.dict_salary[item] / self.dict_quantity[item])
                self.dict_salary_name[item] = math.trunc(self.dict_salary_name[item] / self.dict_quantity_name[item])
            except Exception:
                self.dict_salary_name[item] = 0
                self.dict_quantity_name[item] = 0
                #print(f' vacancie not exist in {item}')  # защита от деления на ноль

        for item in self.dict_salary_city.keys():
            self.dict_salary_city[item] = math.trunc(self.dict_salary_city[item] / self.dict_vacancy_share[item])

        lenght = sum(self.dict_vacancy_share.values())

        for item in self.dict_vacancy_share.keys():
            if self.dict_vacancy_share[item] < (lenght / 100):
                self.dict_vacancy_share[item] = 0
                self.dict_salary_city[item] = 0
            else:
                self.dict_vacancy_share[item] = round(self.dict_vacancy_share[item] / lenght, 4)
        self.dict_vacancy_share = dict(
            sorted(self.dict_vacancy_share.items(), key=lambda item: item[1], reverse=True))

        for item in self.dict_salary_city_name.keys():
            self.dict_salary_city_name[item] = math.trunc(
                self.dict_salary_city_name[item] / self.dict_vacancy_share_name[item])

        lenght = sum(self.dict_vacancy_share_name.values())

        for item in self.dict_vacancy_share_name.keys():
            if self.dict_vacancy_share_name[item] < (lenght / 100):
                self.dict_vacancy_share_name[item] = 0
                self.dict_salary_city_name[item] = 0
            else:
                self.dict_vacancy_share_name[item] = round(self.dict_vacancy_share_name[item] / lenght, 4)
        self.dict_vacancy_share_name = dict(
            sorted(self.dict_vacancy_share_name.items(), key=lambda item: item[1], reverse=True))
        x = {}
        self.dict_salary_city_name = dict(
            sorted(self.dict_salary_city_name.items(), key=lambda item: item[1], reverse=True))
        for i, key in enumerate(self.dict_salary_city_name.keys()):
            if i == 10 or self.dict_salary_city_name[key] == 0:
                break
            x[key] = self.dict_salary_city_name[key]
        # print(f'Уровень зарплат по городам (в порядке убывания): {x}')
        self.dict_salary_city_name = x
        x = {}
        for i, key in enumerate(self.dict_vacancy_share_name.keys()):
            if i == 10 or self.dict_vacancy_share_name[key] == 0:
                break
            x[key] = self.dict_vacancy_share_name[key]
        # print(f'Доля вакансий по городам (в порядке убывания): {x}')
        self.dict_vacancy_share_name = x

        self.dict_salary_city = dict(sorted(self.dict_salary_city.items(), key=lambda item: item[1], reverse=True))
        if not self.dict_salary_name:
            self.dict_salary_name = {2022: 0}
        if not self.dict_quantity_name:
            self.dict_quantity_name = {2022: 0}
        x = {}
        for i, key in enumerate(self.dict_salary_city.keys()):
            if i == 10 or self.dict_salary_city[key] == 0:
                break
            x[key] = self.dict_salary_city[key]
        self.dict_salary_city = x
        x = {}
        for i, key in enumerate(self.dict_vacancy_share.keys()):
            if i == 10 or self.dict_vacancy_share[key] == 0:
                break
            x[key] = self.dict_vacancy_share[key]
        self.dict_vacancy_share = x
        self.dict_salary_name = dict(sorted(self.dict_salary_name.items(), key=lambda item: item[0]))
        self.dict_quantity_name = dict(sorted(self.dict_quantity_name.items(), key=lambda item: item[0]))

    def salaryStat(self, name):
        """
        Собирает статистику и выводит ее через пдф, png,xlsx, текстом в терминал
        Args:
            name(str): название вакансии
        """
        self.__filling_dict(name)
        self.__sort_dict()
        # return [self.dict_salary,
        #         self.dict_quantity,
        #         self.dict_salary_name,
        #         self.dict_quantity_name]
        #
        # print(f'salary = {self.dict_salary}')
        # print(f'quanity =  {self.dict_quantity}')
        # print(f'salary_name =  {self.dict_salary_name}')
        # print(f'quanity_name =  {self.dict_quantity_name}')
        # print(f'city = {self.dict_salary_city}')
        # print(f'share = {self.dict_vacancy_share}')
        # print(f'city_name = {self.dict_salary_city_name}')
        # print(f'share_name = {self.dict_vacancy_share_name}')
        # print(f'name = "{name}"')
        print(f"'{name}':" +
              "{'salary':" +
              f'{self.dict_salary_name},' +
              "'city':" + f'{self.dict_salary_city_name},' + '},'
              )
        # Report(name).generate_excel(
        #     [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name,
        #      self.dict_salary_city, self.dict_vacancy_share])
        #
        img_base64 = Report(name).generate_image(
            [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name,
             self.dict_salary_city, self.dict_vacancy_share])

        Report(name).generate_pdf(
            [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name,
             self.dict_salary_city_name, self.dict_vacancy_share_name],
            img_base64.replace('img ', 'img width="100%"'))


class Report:
    """
    Класс для создания xsls,png,pdf статистик по данным из объекта Statistics
    Atribyties:
        name(str): имя вакансии
        fig(fig): объект для создания гарфика из matplotlib
        ax_1(ax): объект для создания гарфика из matplotlib
        ax_2(ax): объект для создания гарфика из matplotlib
        ax_3(ax): объект для создания гарфика из matplotlib
        ax_4(ax): объект для создания гарфика из matplotlib
    """
    fig, ((ax_1, ax_2), (ax_3, ax_4)) = plt.subplots(nrows=2, ncols=2)

    def __init__(self, name):
        """
        Иницилизация объекта класс Report
        Args:
           name(str): имя вакансии

        >>> type(Report('Програмист')).__name__
        'Report'
        """
        self.name = name

    def generate_excel(self, data):
        """
        Создавние xlsl файла
        Args:
            data(list of dict): данные полученные из объекта Staticstic
        """
        coloumn_1 = {'A': 'Год  ', 'B': 'Средняя зарплата   ', 'C': f'Средняя зарплата - {self.name}    ',
                     'D': 'Количество вакансий  ', 'E': f'Количество вакансий - {self.name}     '}
        coloumn_2 = {'A': 'Город    ', 'B': 'Уровень зарплат    ', 'C': '  ', 'D': 'Город   ',
                     'E': 'Доля вакансий    '}
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet_1 = book.create_sheet("Статистика по годам")
        sheet_2 = book.create_sheet("Статистика по городам")
        sheet_1.append(coloumn_1)
        sheet_2.append(coloumn_2)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for key in data[0].keys():
            sheet_1.append([key, data[0][key], data[2][key], data[1][key], data[3][key]])
        keys = dict(zip(data[4].keys(), data[5].keys()))
        for key in keys.keys():
            sheet_2.append([key, data[4][key], ' ', keys[key], data[5][keys[key]]])

        for key in coloumn_1.keys():
            size = 2
            sheet_1[key + '1'].font = Font(bold=True)
            for i in range(1, len(data[0]) + 2):
                sheet_1[key + str(i)].border = thin_border
                if len(str(sheet_1[key + str(i)].value)) > size:
                    size = len(str(sheet_1[key + str(i)].value))
            sheet_1.column_dimensions[key].width = size

        for key in coloumn_2.keys():
            size = 2
            sheet_2.column_dimensions[key].width = len(coloumn_2[key])
            sheet_2[key + '1'].font = Font(bold=True)
            for i in range(1, len(data[4]) + 2):
                if key == 'E' and i != 1:
                    sheet_2[key + str(i)].number_format = '0.00%'
                if key != 'C':
                    sheet_2[key + str(i)].border = thin_border
                if len(str(sheet_2[key + str(i)].value)) > size:
                    size = len(str(sheet_2[key + str(i)].value))
            sheet_2.column_dimensions[key].width = size

        book.save("report.xlsx")

    def __hist_create(self, data_1, data_2, title, lable_1, lable_2, ax, fig):
        """
        Создание гистограмы
        Args:
            data_1(dict): данные для графика
            data_2(dict): днные для графика
            title(str): заголовок
            lable_1(str): подпись к оси
            lable_2(str): подпись к оси
            ax(ax): объект для создания гарфика из matplotlib
            fig(fig): объект для создания гарфика из matplotlib
        """
        labels = data_1.keys()
        x = np.arange(len(labels))  # the label locations
        width = 0.35  # the width of the bars
        ax.bar(x - width / 2, data_1.values(), width, label=lable_1)
        ax.bar(x + width / 2, data_2.values(), width, label=lable_2)
        ax.axes.get_xaxis().set_ticks([])
        ax.set_title(title)
        ax.tick_params(axis='y', labelsize=8)
        ax.set_xticks(x, labels, fontsize=8, rotation=90)
        ax.legend(fontsize=8)
        ax.yaxis.grid(True)
        fig.tight_layout()
        # plt.show()

    def __horizontal_create(self, data, title, ax):
        """
        Создание гарфика с горизонтальными столбцами
        Args:
            data(dict): данные для графика
            title(str): заголовок
            ax(ax): объект для создания гарфика из matplotlib
        """
        plt.rcdefaults()
        naming_list = [re.sub(re.compile(' |-'), '\n', key) for key in data.keys()]
        city = (naming_list)
        y_pos = np.arange(len(city))
        error = np.random.rand(len(city))
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)
        ax.barh(y_pos, data.values(), xerr=error, align='center')
        ax.set_yticks(y_pos, labels=city)
        ax.invert_yaxis()
        ax.set_title(title)
        ax.xaxis.grid(True)

    def __pie_dreate(self, data, title, ax):
        """
        Создание круговой диаграммы
        Args:
            data(dict): данные для графика
            title(str): заголовок
            ax(ax): объект для создания гарфика из matplotlib
        """
        plt.style.use('_mpl-gallery-nogrid')
        labels = ['Другие'] + list(data.keys())
        x = [1 - sum(data.values())] + list(data.values())
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22',
                  '#17becf', '#6633FF']
        ax.pie(x, labels=labels, colors=colors, shadow=False, startangle=0, textprops={'fontsize': 6}, frame=True)
        ax.set_title(title)
        ax.axis('off')
        ax.axis('equal')

    def generate_image(self, data):
        """
        Создает графики и объединяет их в единый png,
        который  сохраняется в папке проеткта
        Args:
            data(list of dict): данные для графика
        Returns:
            str: картинка в формате base64
        """
        self.__hist_create(data[0], data[2], 'Уровень зарплат по годам', 'средняя з/п', f'з/п {self.name}', self.ax_1,
                           self.fig)
        self.__hist_create(data[1], data[3], 'Количество вакансий по годам', 'Количество вакансий',
                           f'Количество вакансий\n {self.name}', self.ax_2, self.fig)
        self.__horizontal_create(data[4], 'Уровень зарплат по городам', self.ax_3)
        self.__pie_dreate(data[5], 'Доля вакансий по городам', self.ax_4)
        plt.tight_layout()
        plt.savefig('graph.png')
        buf = io.BytesIO()
        self.fig.savefig(buf, format="png")
        data = base64.b64encode(buf.getbuffer()).decode("ascii")
        return f"<img src='data:image/png;base64,{data}'/>"

    def generate_pdf(self, data, img_base64):
        """
        Создает пдф файл из HTML шаблона при помощи Jinjer2
        Args:
            data(list of dict): данные из Staticstic
            img_base64(str): картинка в формате base64
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template(r'pdf_template.html')

        pdf_template = template.render(name=self.name,
                                       dict_salary=data[0],
                                       dict_quantity=data[1],
                                       dict_salary_name=data[2],
                                       dict_quantity_name=data[3],
                                       dict_salary_city=data[4],
                                       dict_vacancy_share=data[5],
                                       image=img_base64)
        with open("F:/Makarov/my_new_file.html", "w",
                  encoding='utf-8-sig') as fh:
            fh.write(pdf_template)
        # config = pdfkit.configuration(wkhtmltopdf=r'F:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        # pdfkit.from_string(pdf_template, r'D:\PycharmProjects\pythonElern\report.pdf', configuration=config)


def main():
    """
    Выполняется при запуске программы
    Ведет диалог с пользователем и вызывает методы для работы с данными
    """
    name = "vacancies_dif_currencies.csv"

    vacanciesDataSet = DataSet(name)
    print('data = {')
    for name_vacancy in vacancies_name.keys():
        if os.stat(name).st_size == 0:
            print("Пустой файл")
        elif len(vacanciesDataSet.getvacancies()) == 0:
            print("Нет данных")
        else:
            # x = input("Введи 1 для вывода таблицы в консоль или 2 для создани PDF\n")
            x = '2'
            if x == '1':
                parameters = input('Введите параметр фильтрации: ')
                sort_info = [input('Введите параметр сортировки: '), input('Обратный порядок сортировки (Да / Нет): ')]
                num = input('Введите диапазон вывода: ')
                name_colomun = input('Введите требуемые столбцы: ')
                num = num.split(' ')
                name_colomun = name_colomun.split(', ')
                if not InputConnect.check_paramtrs(parameters) or not InputConnect.check_sortparam(sort_info):
                    return
                vacanciesDataSet.setvacancies(
                    InputConnect.data_filter(vacanciesDataSet.getvacancies(), parameters, sort_info))
                if vacanciesDataSet.getvacancies():
                    table = InputConnect.create_table(vacanciesDataSet.getvacancies())
                    print(InputConnect.cut_table(table, num, name_colomun))
                else:
                    print('Ничего не найдено')
            elif x == '2':
                stat = Statistics(vacanciesDataSet)
                stat.salaryStat(name_vacancy)
            else:
                print('IllegalArgument')
    print('}')


if __name__ == "__main__":
    main()
