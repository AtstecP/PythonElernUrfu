import base64
import csv
import datetime
import io
import math
import os
import re
import numpy as np
import openpyxl
import matplotlib.pyplot as plt

from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Font, Side, Border
import pdfkit
from prettytable import PrettyTable, ALL

dic_transl = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Нижняя граница вилки оклада',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary': 'Оклад',
    'salary_average': 'Средний оклад',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии(старый формат)',
    'published_at_fixed': 'Дата публикации вакансии',
    '№': '№',
    'noExperience': 'Нет опыта',
    'between1And3': 'От 1 года до 3 лет',
    'between3And6': 'От 3 до 6 лет',
    'moreThan6': 'Более 6 лет',
    'False': 'Нет',
    'True': 'Да',
    'AZN': 'Манаты',
    'BYR': 'Белорусские рубли',
    'EUR': 'Евро',
    'GEL': 'Грузинский лари',
    'KGS': 'Киргизский сом',
    'KZT': 'Тенге',
    'RUR': 'Рубли',
    'UAH': 'Гривны',
    'USD': 'Доллары',
    'UZS': 'Узбекский сум'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:
    def __init__(self, file_name, vacancies_objects):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def getvacancies(self):
        return self.vacancies_objects

    @staticmethod
    def parser_csv(name):
        if os.stat(name).st_size == 0:
            return [], []
        flag = True
        counter = 0
        data = []
        with open(name, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            heading = next(reader)
            # data.append(heading)
            for row in reader:
                for piece in row:
                    if len(piece) == 0:
                        flag = False
                    counter += 1
                if flag and (len(heading) == counter):
                    inf = DataSet.clear_row_csv(row, heading)
                    for key in ['name', 'description', 'key_skills', 'experience_id', 'premium', 'employer_name',
                                'salary', 'area_name',
                                'published_at', 'published_at_fixed']:
                        if not inf.__contains__(key):
                            inf[key] = None
                    data.append(Vacancy(inf['name'],
                                        inf['description'],
                                        inf['key_skills'],
                                        inf['experience_id'],
                                        inf['premium'],
                                        inf['employer_name'],
                                        inf['salary'],
                                        inf['area_name'],
                                        inf['published_at'],
                                        inf['published_at_fixed']))
                flag = True
                counter = 0
        # data.pop(0)
        return data

    @staticmethod
    def formatter(row):
        dict_new = {}
        dict_new['name'] = row['name']
        # dict_new['description'] = row['description']
        # dict_new['key_skills'] = '\n'.join(row['key_skills'])
        # dict_new['experience_id'] = dic_transl[row['experience_id']]
        # dict_new['premium'] = 'Да' if (row['premium'] == 'True') else 'Нет'
        # dict_new['employer_name'] = row['employer_name']
        # salary_from = "{0:,}".format((int(float(row['salary_from'])))).replace(',', ' ')
        # salary_to = "{0:,}".format((int(float(row['salary_to'])))).replace(',', ' ')
        # salary_text = f"{salary_from} - {salary_to} ({dic_transl[row['salary_currency']]}) ({'С вычетом' if row['salary_gross'] == 'False' else 'Без вычета'} налогов)"
        salary_text = None
        row['salary_gross'] = None
        dict_new['salary'] = Salary(salary_text, row["salary_from"], row["salary_to"], row['salary_gross'],
                                    row['salary_currency'], math.floor(
                (int(row['salary_from'].replace('.0', '')) + int(row['salary_to'].replace('.0', ''))) / 2))
        dict_new['area_name'] = row['area_name']
        dict_new['published_at'] = datetime.datetime.strptime(row['published_at'].replace('T', ' '),
                                                              '%Y-%m-%d %H:%M:%S+%f')
        dict_new[
            'published_at_fixed'] = f"{row['published_at'][8:10]}.{row['published_at'][5:7]}.{row['published_at'][:4]}"
        return dict_new

    @staticmethod
    def clear_row_csv(row, list_naming):
        # list_naming = ['name', 'description', 'key_skills', 'experience_id', 'premium', 'employer_name', 'salary', 'area_name',
        #         'published_at', 'published_at_fixed']
        dict_new = {}
        pattern = re.compile('<.*?>')
        pattern1 = re.compile('\s+')
        for i in range(0, len(list_naming)):
            if list_naming[i] == 'key_skills':
                dict_new[list_naming[i]] = row[i].split("\n")
            else:
                # row[i] = row[i].replace("\n", ", ")
                row[i] = re.sub(pattern, '', row[i])
                row[i] = re.sub(pattern1, ' ', row[i])
                row[i] = row[i].strip()
                dict_new[list_naming[i]] = row[i]

        return DataSet.formatter(dict_new)


class Vacancy:
    def __init__(self, name, description, key_skills, experience_id, premium, employer_name, salary, area_name,
                 published_at, published_at_fixed):
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at
        self.published_at_fixed = published_at_fixed

    def getdata(self):
        return [self.name, self.description, self.key_skills, self.experience_id, self.premium, self.employer_name,
                self.salary.salary_text, self.area_name, self.published_at, self.published_at_fixed]


class Salary:
    def __init__(self, salary_text, salary_from, salary_to, salary_gross, salary_currency, salary_average):
        self.salary_text = salary_text
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency
        self.salary_average = salary_average


class InputConnect:

    @staticmethod
    def check_paramtrs(parameters):
        if len(parameters) == 0:
            return True
        parameters_list = parameters.split(': ')
        if ':' not in parameters:
            print('Формат ввода некорректен')
            return False
        if parameters[:parameters.index(':')] not in dic_transl.values():
            print('Параметр поиска некорректен')
            return False
        parameters_list[0] = ''.join(filter(lambda x: parameters_list[0] == dic_transl[x], dic_transl))
        return bool(parameters_list[0])

    @staticmethod
    def check_sortparam(sort_info):
        if (sort_info[0] not in dic_transl.values()) and (sort_info[0] != ''):
            print('Параметр сортировки некорректен')
            return False
        if (sort_info[1] != 'Да') and (sort_info[1] != 'Нет') and (sort_info[1] != ''):
            print('Порядок сортировки задан некорректно')
            return False
        return True

    @staticmethod
    def sort_vacancies(vacancies, sort_info):
        vacancies_sort = vacancies
        if sort_info[0] == 'Оклад':
            data_sort = sorted(vacancies_sort,
                               key=lambda x: x.salary.salary_average * currency_to_rub[x.salary.salary_currency],
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Навыки':
            data_sort = sorted(vacancies_sort, key=lambda x: x.key_skills.count('\n'),
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Опыт работы':
            experience_weight = {'Нет опыта': 0, 'От 1 года до 3 лет': 1, 'От 3 до 6 лет': 2, 'Более 6 лет': 3}
            data_sort = sorted(vacancies_sort, key=lambda x: experience_weight[x.experience_id],
                               reverse=True if sort_info[1] == 'Да' else False)
        elif sort_info[0] == 'Дата публикации вакансии':
            data_sort = sorted(vacancies_sort, key=lambda x: x.published_at,
                               reverse=True if sort_info[1] == 'Да' else False)
        else:
            data_sort = sorted(vacancies_sort,
                               key=lambda x: getattr(x, ''.join(
                                   filter(lambda x: sort_info[0] == dic_transl[x], dic_transl))),
                               reverse=True if sort_info[1] == 'Да' else False)
        return data_sort

    @staticmethod
    def data_filter(vacancies, parameters, sort_info):
        if len(parameters) == 0:
            if sort_info[0] != '':
                return InputConnect.sort_vacancies(vacancies, sort_info)
            return vacancies
        parameters_list = parameters.split(': ')
        data = []
        parameters_list[0] = ''.join(filter(lambda x: parameters_list[0] == dic_transl[x], dic_transl))
        i = ''.join(filter(lambda x: parameters_list[1] == dic_transl[x], dic_transl))
        if i and parameters_list[0] not in ['experience_id', 'premium']:
            parameters_list[1] = i
        if parameters_list[0] == 'key_skills':
            parameters_list[1] = parameters_list[1].split(', ')
            for vacancy in vacancies:
                f = True
                for parametr in parameters_list[1]:
                    if (parametr + '\n') not in vacancy.key_skills:
                        f = False
                if f:
                    data.append(vacancy)
        elif parameters_list[0] == 'salary':
            for vacancy in vacancies:
                if float(vacancy.salary.salary_from) <= float(parameters_list[1]) <= float(vacancy.salary.salary_to):
                    data.append(vacancy)
        # elif parameters_list[0] == 'published_at' or parameters_list[0] == 'published_at_fixed':
        elif parameters_list[0] == 'published_at_fixed':
            for vacancy in vacancies:
                if vacancy.published_at_fixed == parameters_list[1]:
                    data.append(vacancy)
        elif parameters_list[0] == 'salary_currency':
            for vacancy in vacancies:
                if parameters_list[1] == vacancy.salary.salary_currency:
                    data.append(vacancy)
        else:
            for vacancy in vacancies:
                if parameters_list[1] == getattr(vacancy, parameters_list[0]):
                    data.append(vacancy)
        if sort_info[0] != '':
            data = InputConnect.sort_vacancies(data, sort_info)
        return data

    @staticmethod
    def cut_table(table, num_str, name_colm):
        if len(num_str) == 1:
            return table.get_string(start=0 if num_str[0] == '' else int(num_str[0]) - 1,
                                    fields=table.field_names if name_colm[0] == '' else ['№'] + name_colm)
        else:
            return table.get_string(start=int(num_str[0]) - 1, end=int(num_str[1]) - 1,
                                    fields=table.field_names if name_colm[0] == '' else ['№'] + name_colm)

    @staticmethod
    def translate(list_naming):
        list_translate = []
        for name in list_naming:
            list_translate.append(dic_transl[name])
        return list_translate

    @staticmethod
    def create_table(vacancies):
        data = []
        for vacancy in vacancies:
            data.append(vacancy)
        table = PrettyTable()
        table.field_names = ['№'] + InputConnect.translate(list(data[0].__dict__))
        table.hrules = ALL
        table.align = 'l'
        table.max_width = 20
        for i, row in enumerate(data):
            # row['key_skills'] = reduce(lambda x, y: x + '\n' + y, row['key_skills'])
            if len(row.key_skills) > 100:
                row.key_skills = row.key_skills[:100] + '...'
            if len(row.description) > 100:
                row.description = row.description[:100] + '...'
            table.add_row([i + 1] + row.getdata())
        return table


class Statistics:
    def __init__(self, dataset):
        self.dataset = dataset
        self.dict_salary = {}
        self.dict_quantity = {}
        self.dict_salary_name = {}
        self.dict_quantity_name = {}
        self.dict_salary_city = {}
        self.dict_vacancy_share = {}

    def __filling_dict(self, name):
        for vacancy in self.dataset.getvacancies():
            if not self.dict_salary.__contains__(vacancy.published_at.year):
                self.dict_salary[vacancy.published_at.year] = vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_quantity[vacancy.published_at.year] = 1
            else:
                self.dict_salary[vacancy.published_at.year] += vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_quantity[vacancy.published_at.year] += 1

            if not self.dict_salary_city.__contains__(vacancy.area_name):
                self.dict_salary_city[vacancy.area_name] = vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_vacancy_share[vacancy.area_name] = 1
            else:
                self.dict_salary_city[vacancy.area_name] += vacancy.salary.salary_average * currency_to_rub[
                    vacancy.salary.salary_currency]
                self.dict_vacancy_share[vacancy.area_name] += 1

            if name in vacancy.name:
                if not self.dict_salary_name.__contains__(vacancy.published_at.year):
                    self.dict_salary_name[vacancy.published_at.year] = vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_quantity_name[vacancy.published_at.year] = 1
                else:
                    self.dict_salary_name[vacancy.published_at.year] += vacancy.salary.salary_average * currency_to_rub[
                        vacancy.salary.salary_currency]
                    self.dict_quantity_name[vacancy.published_at.year] += 1

    def __sort_dict(self):
        for item in self.dict_salary.keys():
            self.dict_salary[item] = math.trunc(self.dict_salary[item] / self.dict_quantity[item])
        for item in self.dict_salary_name.keys():
            self.dict_salary_name[item] = math.trunc(self.dict_salary_name[item] / self.dict_quantity_name[item])
        for item in self.dict_salary_city.keys():
            self.dict_salary_city[item] = math.trunc(self.dict_salary_city[item] / self.dict_vacancy_share[item])

        lenght = sum(self.dict_vacancy_share.values())

        for item in self.dict_vacancy_share.keys():
            if self.dict_vacancy_share[item] < (lenght / 100):
                self.dict_vacancy_share[item] = 0
                self.dict_salary_city[item] = 0
            else:
                self.dict_vacancy_share[item] = round(self.dict_vacancy_share[item] / lenght, 4)
        self.dict_vacancy_share = dict(
            sorted(self.dict_vacancy_share.items(), key=lambda item: item[1], reverse=True))
        self.dict_salary_city = dict(sorted(self.dict_salary_city.items(), key=lambda item: item[1], reverse=True))
        if not self.dict_salary_name:
            self.dict_salary_name = {2022: 0}
        if not self.dict_quantity_name:
            self.dict_quantity_name = {2022: 0}
        # print(f'Динамика уровня зарплат по годам: {dict_salary}')
        # print(f'Динамика количества вакансий по годам: {dict_quantity}')
        # print(f'Динамика уровня зарплат по годам для выбранной профессии: {dict_salary_name}')
        # print(f'Динамика количества вакансий по годам для выбранной профессии: {dict_quantity_name}')
        x = {}
        for i, key in enumerate(self.dict_salary_city.keys()):
            if i == 10 or self.dict_salary_city[key] == 0:
                break
            x[key] = self.dict_salary_city[key]
        # print(f'Уровень зарплат по городам (в порядке убывания): {x}')
        self.dict_salary_city = x
        x = {}
        for i, key in enumerate(self.dict_vacancy_share.keys()):
            if i == 10 or self.dict_vacancy_share[key] == 0:
                break
            x[key] = self.dict_vacancy_share[key]
        # print(f'Доля вакансий по городам (в порядке убывания): {x}')
        self.dict_vacancy_share = x

    def salaryStat(self, name):
        self.__filling_dict(name)
        self.__sort_dict()

        Report(name).generate_excel(
            [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name, self.dict_salary_city, self.dict_vacancy_share])
        img_base64 = Report(name).generate_image(
            [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name,
             self.dict_salary_city, self.dict_vacancy_share])
        Report(name).generate_pdf(
            [self.dict_salary, self.dict_quantity, self.dict_salary_name, self.dict_quantity_name,
             self.dict_salary_city, self.dict_vacancy_share],
            img_base64.replace('img ', 'img width="100%"'))


class Report:
    fig, ((ax_1, ax_2), (ax_3, ax_4)) = plt.subplots(nrows=2, ncols=2)

    def __init__(self, name):
        self.name = name

    def generate_excel(self, data):
        coloumn_1 = {'A': 'Год  ', 'B': 'Средняя зарплата   ', 'C': f'Средняя зарплата - {self.name}    ',
                     'D': 'Количество вакансий  ', 'E': f'Количество вакансий - {self.name}     '}
        coloumn_2 = {'A': 'Город    ', 'B': 'Уровень зарплат    ', 'C': '  ', 'D': 'Город   ',
                     'E': 'Доля вакансий    '}
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet_1 = book.create_sheet("Статистика по годам")
        sheet_2 = book.create_sheet("Статистика по городам")
        sheet_1.append(coloumn_1)
        sheet_2.append(coloumn_2)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for key in data[0].keys():
            sheet_1.append([key, data[0][key], data[2][key], data[1][key], data[3][key]])
        keys = dict(zip(data[4].keys(), data[5].keys()))
        for key in keys.keys():
            sheet_2.append([key, data[4][key], ' ', keys[key], data[5][keys[key]]])

        for key in coloumn_1.keys():
            size = 2
            sheet_1[key + '1'].font = Font(bold=True)
            for i in range(1, len(data[0]) + 2):
                sheet_1[key + str(i)].border = thin_border
                if len(str(sheet_1[key + str(i)].value)) > size:
                    size = len(str(sheet_1[key + str(i)].value))
            sheet_1.column_dimensions[key].width = size

        for key in coloumn_2.keys():
            size = 2
            sheet_2.column_dimensions[key].width = len(coloumn_2[key])
            sheet_2[key + '1'].font = Font(bold=True)
            for i in range(1, len(data[4]) + 2):
                if key == 'E' and i != 1:
                    sheet_2[key + str(i)].number_format = '0.00%'
                if key != 'C':
                    sheet_2[key + str(i)].border = thin_border
                if len(str(sheet_2[key + str(i)].value)) > size:
                    size = len(str(sheet_2[key + str(i)].value))
            sheet_2.column_dimensions[key].width = size

        book.save("report.xlsx")

    def __hist_create(self, data_1, data_2, title, lable_1, lable_2, ax, fig):
        labels = data_1.keys()
        x = np.arange(len(labels))  # the label locations
        width = 0.35  # the width of the bars
        ax.bar(x - width / 2, data_1.values(), width, label=lable_1)
        ax.bar(x + width / 2, data_2.values(), width, label=lable_2)
        ax.axes.get_xaxis().set_ticks([])
        ax.set_title(title)
        ax.tick_params(axis='y', labelsize=8)
        ax.set_xticks(x, labels, fontsize=8, rotation=90)
        ax.legend(fontsize=8)
        ax.yaxis.grid(True)
        fig.tight_layout()
        # plt.show()

    def __horizontal_create(self, data, data1, title, ax):
        plt.rcdefaults()
        naming_list = [re.sub(re.compile(' |-'), '\n', key) for key in data.keys()]
        city = (naming_list)
        y_pos = np.arange(len(city))
        error = np.random.rand(len(city))
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)
        ax.barh(y_pos, data.values(), xerr=error, align='center')
        ax.set_yticks(y_pos, labels=city)
        ax.invert_yaxis()
        ax.set_title(title)
        ax.xaxis.grid(True)

    def __pie_dreate(self, data, title, ax):
        plt.style.use('_mpl-gallery-nogrid')
        labels = ['Другие'] + list(data.keys())
        x = [1 - sum(data.values())] + list(data.values())
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22',
                  '#17becf', '#6633FF']
        ax.pie(x, labels=labels, colors=colors, shadow=False, startangle=0, textprops={'fontsize': 6}, frame=True)
        ax.set_title(title)
        ax.axis('off')
        ax.axis('equal')

    def generate_image(self, data):
        self.__hist_create(data[0], data[2], 'Уровень зарплат по годам', 'средняя з/п', f'з/п {self.name}', self.ax_1,
                           self.fig)
        self.__hist_create(data[1], data[3], 'Количество вакансий по годам', 'Количество вакансий',
                           f'Количество вакансий\n {self.name}', self.ax_2, self.fig)
        self.__horizontal_create(data[4], data[5], 'Уровень зарплат по городам', self.ax_3)
        self.__pie_dreate(data[5], 'Доля вакансий по городам', self.ax_4)
        plt.tight_layout()
        plt.savefig('graph.png')
        buf = io.BytesIO()
        self.fig.savefig(buf, format="png")
        data = base64.b64encode(buf.getbuffer()).decode("ascii")
        return f"<img src='data:image/png;base64,{data}'/>"

    def generate_pdf(self, data, img_base64):
        env = Environment(loader=FileSystemLoader('..'))
        template = env.get_template(r'pdf_template.html')  # absolute path to html

        pdf_template = template.render(name=self.name,
                                       dict_salary=data[0],
                                       dict_quantity=data[1],
                                       dict_salary_name=data[2],
                                       dict_quantity_name=data[3],
                                       dict_salary_city=data[4],
                                       dict_vacancy_share=data[5],
                                       image=img_base64)

        config = pdfkit.configuration(wkhtmltopdf=r'F:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, r'F:\PycharmProjects\pythonElern\report.pdf', configuration=config)


# {'dict_quantity': data[1]},'dict_salary_name': data[3]}, {'dict_quantity_name': data[4]},

def main():
    # data = [{2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658,
    #          2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335, 2019: 69467, 2020: 73431, 2021: 82690, 2022: 91795}
    #     , {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837,
    #        2015: 70039, 2016: 75145, 2017: 82823, 2018: 131701, 2019: 115086, 2020: 102243, 2021: 57623, 2022: 18294}
    #     , {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384,
    #        2015: 62322, 2016: 66817, 2017: 72460, 2018: 76879, 2019: 85300, 2020: 89791, 2021: 100987, 2022: 116651}
    #     , {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492, 2015: 5375,
    #        2016: 7219, 2017: 8105, 2018: 10062, 2019: 9016, 2020: 7113, 2021: 3466, 2022: 1115}
    #     , {'Москва': 76970, 'Санкт-Петербург': 65286, 'Новосибирск': 62254, 'Екатеринбург': 60962, 'Казань': 52580,
    #        'Краснодар': 51644, 'Челябинск': 51265, 'Самара': 50994, 'Пермь': 48089, 'Нижний Новгород': 47662}
    #     , {'Москва': 0.3246, 'Санкт-Петербург': 0.1197, 'Новосибирск': 0.0271, 'Казань': 0.0237,
    #        'Нижний Новгород': 0.0232, 'Ростов-на-Дону': 0.0209, 'Екатеринбург': 0.0207, 'Краснодар': 0.0185,
    #        'Самара': 0.0143, 'Воронеж': 0.0141}]
    #
    # name = input('Введите название файла: ')
    name = "vacancies_by_year.csv"
    # name_vacancy = input('Введите название профессии: ')
    name_vacancy = 'Программист'
    vacanciesDataSet = DataSet(name, DataSet.parser_csv(name))
    if os.stat(name).st_size == 0:
        print("Пустой файл")
    elif len(vacanciesDataSet.getvacancies()) == 0:
        print("Нет данных")
    else:
        stat = Statistics(vacanciesDataSet)
        stat.salaryStat(name_vacancy)


if __name__ == "__main__":
    main()
